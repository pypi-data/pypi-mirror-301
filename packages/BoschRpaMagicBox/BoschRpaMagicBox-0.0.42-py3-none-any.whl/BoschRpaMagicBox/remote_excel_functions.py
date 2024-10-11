from io import BytesIO
import pandas as pd
from openpyxl import load_workbook
from helper_functions import smb_load_file_obj, smb_store_remote_file_by_obj


def save_dataframe_into_excel(username, password, server_ip, server_name, share_name, remote_file_path, sheet_name, saved_data, start_cell_location='A1',
                              keep_header=True, port=445):
    """ This function is used to save data in Dataframe format to an Excel file.

    Args:

        username(str): This is the username
        password(str): This is the password
        server_ip(str): This is the ip address of the public folder, which can be same as server name
        server_name(str):This is the server name (url), e.g. szh0fs06.apac.bosch.com
        share_name(str): This is the share_name of public folder, e.g. GS_ACC_CN$
        remote_file_path(str): This is the public file path that file will be saved under share name folder
        port(int): This is the port number of the server name
        sheet_name(str): The name of the sheet.
        keep_header(bool): Whether to keep the header of the data.
        start_cell_location(str): The location of start cell. e.g. A1
        saved_data(pd.DataFrame): The data to be saved.
    """
    file_obj = smb_load_file_obj(username, password, server_ip, server_name, share_name, remote_file_path, port)
    target_workbook = load_workbook(file_obj)

    target_sheet = target_workbook[sheet_name]
    start_cell = target_sheet[start_cell_location]
    start_cell_row = start_cell.row
    start_cell_column = start_cell.column

    data_columns = list(saved_data.columns)

    if keep_header:
        start_column_number = start_cell_column
        for column_index, column_name in enumerate(data_columns):
            target_sheet.cell(row=start_cell_row, column=start_column_number + column_index, value=column_name)
        start_cell_row += 1

    for column_name in data_columns:
        start_row_number = start_cell_row
        for column_value in saved_data[column_name]:
            target_sheet.cell(row=start_row_number, column=start_cell_column, value=column_value)
            start_row_number += 1
        start_cell_column += 1

    output_io = BytesIO()
    target_workbook.save(output_io)
    output_io.seek(0)

    smb_store_remote_file_by_obj(username, password, server_ip, server_name, share_name, remote_file_path, output_io, port)


def save_series_into_excel(username, password, server_ip, server_name, share_name, remote_file_path, sheet_name, saved_data, start_cell_location='A1',
                           port=445):
    """ This function is used to save data in Series or list format to an Excel file.

    Args:

        username(str): This is the username
        password(str): This is the password
        server_ip(str): This is the ip address of the public folder, which can be same as server name
        server_name(str):This is the server name (url), e.g. szh0fs06.apac.bosch.com
        share_name(str): This is the share_name of public folder, e.g. GS_ACC_CN$
        remote_file_path(str): This is the public file path that file will be saved under share name folder
        port(int): This is the port number of the server name
        sheet_name(str): The name of the sheet.
        start_cell_location(str): The location of start cell. e.g. A1
        saved_data(pd.Series | list): The data to be saved.
    """
    file_obj = smb_load_file_obj(username, password, server_ip, server_name, share_name, remote_file_path, port)
    target_workbook = load_workbook(file_obj)

    target_sheet = target_workbook[sheet_name]
    start_cell = target_sheet[start_cell_location]
    start_cell_row = start_cell.row
    start_cell_column = start_cell.column

    for save_value in saved_data:
        target_sheet.cell(row=start_cell_row, column=start_cell_column, value=save_value)
        start_cell_row += 1

    output_io = BytesIO()
    target_workbook.save(output_io)
    output_io.seek(0)

    smb_store_remote_file_by_obj(username, password, server_ip, server_name, share_name, remote_file_path, output_io, port)


def save_single_value_into_excel(username, password, server_ip, server_name, share_name, remote_file_path, sheet_name, saved_data, target_cell_location='A1',
                                 port=445):
    """ This function is used to save single value to an Excel file.

    Args:

        username(str): This is the username
        password(str): This is the password
        server_ip(str): This is the ip address of the public folder, which can be same as server name
        server_name(str):This is the server name (url), e.g. szh0fs06.apac.bosch.com
        share_name(str): This is the share_name of public folder, e.g. GS_ACC_CN$
        remote_file_path(str): This is the public file path that file will be saved under share name folder
        port(int): This is the port number of the server name
        sheet_name(str): The name of the sheet.
        target_cell_location(str): The location of target cell. e.g. A1
        saved_data(str|int|float): The data to be saved.
    """
    file_obj = smb_load_file_obj(username, password, server_ip, server_name, share_name, remote_file_path, port)
    target_workbook = load_workbook(file_obj)

    target_sheet = target_workbook[sheet_name]
    target_sheet[target_cell_location] = saved_data

    output_io = BytesIO()
    target_workbook.save(output_io)
    output_io.seek(0)

    smb_store_remote_file_by_obj(username, password, server_ip, server_name, share_name, remote_file_path, output_io, port)


def save_multiple_value_into_excel(username, password, server_ip, server_name, share_name, remote_file_path, sheet_name, saved_data, port=445):
    """ This function is used to save multiple values to an Excel file.

    Args:

        username(str): This is the username
        password(str): This is the password
        server_ip(str): This is the ip address of the public folder, which can be same as server name
        server_name(str):This is the server name (url), e.g. szh0fs06.apac.bosch.com
        share_name(str): This is the share_name of public folder, e.g. GS_ACC_CN$
        remote_file_path(str): This is the public file path that file will be saved under share name folder
        port(int): This is the port number of the server name
        sheet_name(str): The name of the sheet.
        saved_data(dict): The data to be saved.e.g. {'A1': 1, 'B2': 2}
    """
    file_obj = smb_load_file_obj(username, password, server_ip, server_name, share_name, remote_file_path, port)
    target_workbook = load_workbook(file_obj)

    target_sheet = target_workbook[sheet_name]

    for cell_location, cell_value in saved_data.items():
        target_sheet[cell_location] = cell_value

    output_io = BytesIO()
    target_workbook.save(output_io)
    output_io.seek(0)

    smb_store_remote_file_by_obj(username, password, server_ip, server_name, share_name, remote_file_path, output_io, port)


def append_flexible_dataframe_into_excel(username, password, server_ip, server_name, share_name, remote_file_path, sheet_name, saved_data, column_name_dict,
                                         port=445):
    """ This function is used to append data in DataFrame format to an Excel file.

    Args:

        username(str): This is the username
        password(str): This is the password
        server_ip(str): This is the ip address of the public folder, which can be same as server name
        server_name(str):This is the server name (url), e.g. szh0fs06.apac.bosch.com
        share_name(str): This is the share_name of public folder, e.g. GS_ACC_CN$
        remote_file_path(str): This is the public file path that file will be saved under share name folder
        port(int): This is the port number of the server name
        sheet_name(str): The name of the sheet.
        saved_data(pd.Dataframe): The data to be saved.
        column_name_dict(dict): The dictionary of column name and column index. e.g. {'name':'A', 'age':'C'}
    """
    file_obj = smb_load_file_obj(username, password, server_ip, server_name, share_name, remote_file_path, port)
    target_workbook = load_workbook(file_obj)

    target_sheet = target_workbook[sheet_name]
    max_row_number = target_sheet.max_row

    for column_name, column_index in column_name_dict.items():
        start_row_number = max_row_number + 1
        for column_value in saved_data[column_name]:
            target_sheet[f'{column_index}{start_row_number}'] = column_value
            start_row_number += 1

    output_io = BytesIO()
    target_workbook.save(output_io)
    output_io.seek(0)

    smb_store_remote_file_by_obj(username, password, server_ip, server_name, share_name, remote_file_path, output_io, port)
