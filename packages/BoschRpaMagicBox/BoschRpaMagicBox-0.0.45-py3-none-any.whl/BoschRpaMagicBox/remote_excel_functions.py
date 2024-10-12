from io import BytesIO
import pandas as pd
from openpyxl import load_workbook, Workbook
from helper_functions import smb_load_file_obj, smb_store_remote_file_by_obj


def create_excel_file(username, password, server_ip, server_name, share_name, remote_file_path, sheet_name='Sheet1', port=445):
    """ This function is used to create an Excel file.

    Args:

        username(str): This is the username
        password(str): This is the password
        server_ip(str): This is the ip address of the public folder, which can be same as server name
        server_name(str):This is the server name (url), e.g. szh0fs06.apac.bosch.com
        share_name(str): This is the share_name of public folder, e.g. GS_ACC_CN$
        remote_file_path(str): This is the public file path that file will be saved under share name folder
        port(int): This is the port number of the server name
        sheet_name(str): The name of the sheet.
    """
    target_workbook = Workbook()
    target_sheet = target_workbook.active
    target_sheet.title = sheet_name
    output_io = BytesIO()
    target_workbook.save(output_io)
    output_io.seek(0)

    smb_store_remote_file_by_obj(username, password, server_ip, server_name, share_name, remote_file_path, output_io, port)


def check_sheet_exists(workbook, sheet_name):
    """ This function is used to check whether the sheet exists in

    Args:
        workbook(Workbook): This is the instance of Workbook
        sheet_name(str): The name of the sheet.

    """
    return sheet_name in workbook.sheetnames


def add_column_filter(worksheet):
    """ This function is used to add column filter in the worksheet.

    Args:
        worksheet(Worksheet): This is the instance of Worksheet

    """
    worksheet.auto_filter.ref = worksheet.dimensions
    return worksheet


def auto_set_column_width(worksheet):
    """ This function is used to auto set column width in the worksheet.

    Args:
        worksheet(Worksheet): This is the instance of Worksheet

    """
    for column in worksheet.columns:
        max_length = 0
        column = column[0].column_letter  # 获取列字母
        for cell in column:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2  # 加点余量
        worksheet.column_dimensions[column].width = adjusted_width

    return worksheet


def manage_workbook_sheet(username, password, server_ip, server_name, share_name, remote_file_path, rename_sheet_dict=None, add_sheet_name_list=None, delete_sheet_name_list=None,
                          port=445):
    """

    Args:
        username(str): This is the username
        password(str): This is the password
        server_ip(str): This is the ip address of the public folder, which can be same as server name
        server_name(str):This is the server name (url), e.g. szh0fs06.apac.bosch.com
        share_name(str): This is the share_name of public folder, e.g. GS_ACC_CN$
        remote_file_path(str): This is the public file path that file will be saved under share name folder
        port(int): This is the port number of the server name
        rename_sheet_dict(dict): {old_sheet_name: new_sheet_name}
        add_sheet_name_list(list):This is the list of sheet that will be created
        delete_sheet_name_list(list): This is the list of sheet that will be deleted

    """
    if rename_sheet_dict is None:
        rename_sheet_dict = {}
    if add_sheet_name_list is None:
        add_sheet_name_list = []
    if delete_sheet_name_list is None:
        delete_sheet_name_list = []

    file_obj = smb_load_file_obj(username, password, server_ip, server_name, share_name, remote_file_path, port)
    target_workbook = load_workbook(file_obj)

    for old_sheet_name, new_sheet_name in rename_sheet_dict.items():
        if check_sheet_exists(target_workbook, old_sheet_name):
            target_workbook[old_sheet_name].title = new_sheet_name

    for add_sheet_name in add_sheet_name_list:
        if not check_sheet_exists(target_workbook, add_sheet_name):
            target_workbook.create_sheet(add_sheet_name)

    for delete_sheet_name in delete_sheet_name_list:
        if check_sheet_exists(target_workbook, delete_sheet_name):
            target_workbook.remove(target_workbook[delete_sheet_name])

    output_io = BytesIO()
    target_workbook.save(output_io)
    output_io.seek(0)

    smb_store_remote_file_by_obj(username, password, server_ip, server_name, share_name, remote_file_path, output_io, port)


def save_dataframe_into_excel(username, password, server_ip, server_name, share_name, remote_file_path, sheet_name, saved_data, start_cell_location='A1',
                              keep_header=True, port=445, auto_filter=False, auto_column_width=False):
    """ This function is used to save data in Dataframe format to an Excel file.

    Args:
        username(str): This is the username
        password(str): This is the password
        server_ip(str): This is the ip address of the public folder, which can be same as server name
        server_name(str):This is the server name (url), e.g. szh0fs06.apac.bosch.com
        share_name(str): This is the share_name of public folder, e.g. GS_ACC_CN$
        remote_file_path(str): This is the public file path that file will be saved under share name folder
        port(int): This is the port number of the server name
        sheet_name(str): The name of the sheet.
        keep_header(bool): Whether to keep the header of the data.
        start_cell_location(str): The location of start cell. e.g. A1
        saved_data(pd.DataFrame): The data to be saved.
        auto_filter(bool): Whether to add column filter.
        auto_column_width(bool): Whether to auto set column width.
    """
    file_obj = smb_load_file_obj(username, password, server_ip, server_name, share_name, remote_file_path, port)
    target_workbook = load_workbook(file_obj)

    target_sheet = target_workbook[sheet_name]
    start_cell = target_sheet[start_cell_location]
    start_cell_row = start_cell.row
    start_cell_column = start_cell.column

    data_columns = list(saved_data.columns)

    if keep_header:
        start_column_number = start_cell_column
        for column_index, column_name in enumerate(data_columns):
            target_sheet.cell(row=start_cell_row, column=start_column_number + column_index, value=column_name)
        start_cell_row += 1

    for column_name in data_columns:
        start_row_number = start_cell_row
        for column_value in saved_data[column_name]:
            target_sheet.cell(row=start_row_number, column=start_cell_column, value=column_value)
            start_row_number += 1
        start_cell_column += 1

    if auto_filter:
        target_sheet = add_column_filter(target_sheet)

    if auto_column_width:
        auto_set_column_width(target_sheet)

    output_io = BytesIO()
    target_workbook.save(output_io)
    output_io.seek(0)

    smb_store_remote_file_by_obj(username, password, server_ip, server_name, share_name, remote_file_path, output_io, port)


def save_series_into_excel(username, password, server_ip, server_name, share_name, remote_file_path, sheet_name, saved_data, start_cell_location='A1',
                           port=445, auto_filter=False, auto_column_width=False):
    """ This function is used to save data in Series or list format to an Excel file.

    Args:

        username(str): This is the username
        password(str): This is the password
        server_ip(str): This is the ip address of the public folder, which can be same as server name
        server_name(str):This is the server name (url), e.g. szh0fs06.apac.bosch.com
        share_name(str): This is the share_name of public folder, e.g. GS_ACC_CN$
        remote_file_path(str): This is the public file path that file will be saved under share name folder
        port(int): This is the port number of the server name
        sheet_name(str): The name of the sheet.
        start_cell_location(str): The location of start cell. e.g. A1
        saved_data(pd.Series | list): The data to be saved.
        auto_filter(bool): Whether to add column filter.
        auto_column_width(bool): Whether to auto set column width.
    """
    file_obj = smb_load_file_obj(username, password, server_ip, server_name, share_name, remote_file_path, port)
    target_workbook = load_workbook(file_obj)

    target_sheet = target_workbook[sheet_name]
    start_cell = target_sheet[start_cell_location]
    start_cell_row = start_cell.row
    start_cell_column = start_cell.column

    for save_value in saved_data:
        target_sheet.cell(row=start_cell_row, column=start_cell_column, value=save_value)
        start_cell_row += 1

    if auto_filter:
        target_sheet = add_column_filter(target_sheet)

    if auto_column_width:
        auto_set_column_width(target_sheet)

    output_io = BytesIO()
    target_workbook.save(output_io)
    output_io.seek(0)

    smb_store_remote_file_by_obj(username, password, server_ip, server_name, share_name, remote_file_path, output_io, port)


def save_single_value_into_excel(username, password, server_ip, server_name, share_name, remote_file_path, sheet_name, saved_data, target_cell_location='A1',
                                 port=445, auto_filter=False, auto_column_width=False):
    """ This function is used to save single value to an Excel file.

    Args:

        username(str): This is the username
        password(str): This is the password
        server_ip(str): This is the ip address of the public folder, which can be same as server name
        server_name(str):This is the server name (url), e.g. szh0fs06.apac.bosch.com
        share_name(str): This is the share_name of public folder, e.g. GS_ACC_CN$
        remote_file_path(str): This is the public file path that file will be saved under share name folder
        port(int): This is the port number of the server name
        sheet_name(str): The name of the sheet.
        target_cell_location(str): The location of target cell. e.g. A1
        saved_data(str|int|float): The data to be saved.
        auto_filter(bool): Whether to add column filter.
        auto_column_width(bool): Whether to auto set column width.
    """
    file_obj = smb_load_file_obj(username, password, server_ip, server_name, share_name, remote_file_path, port)
    target_workbook = load_workbook(file_obj)

    target_sheet = target_workbook[sheet_name]
    target_sheet[target_cell_location] = saved_data

    if auto_filter:
        target_sheet = add_column_filter(target_sheet)

    if auto_column_width:
        auto_set_column_width(target_sheet)

    output_io = BytesIO()
    target_workbook.save(output_io)
    output_io.seek(0)

    smb_store_remote_file_by_obj(username, password, server_ip, server_name, share_name, remote_file_path, output_io, port)


def save_multiple_value_into_excel(username, password, server_ip, server_name, share_name, remote_file_path, sheet_name, saved_data, port=445, auto_filter=False,
                                   auto_column_width=False):
    """ This function is used to save multiple values to an Excel file.

    Args:

        username(str): This is the username
        password(str): This is the password
        server_ip(str): This is the ip address of the public folder, which can be same as server name
        server_name(str):This is the server name (url), e.g. szh0fs06.apac.bosch.com
        share_name(str): This is the share_name of public folder, e.g. GS_ACC_CN$
        remote_file_path(str): This is the public file path that file will be saved under share name folder
        port(int): This is the port number of the server name
        sheet_name(str): The name of the sheet.
        saved_data(dict): The data to be saved.e.g. {'A1': 1, 'B2': 2}
        auto_filter(bool): Whether to add column filter.
        auto_column_width(bool): Whether to auto set column width.
    """
    file_obj = smb_load_file_obj(username, password, server_ip, server_name, share_name, remote_file_path, port)
    target_workbook = load_workbook(file_obj)

    target_sheet = target_workbook[sheet_name]

    for cell_location, cell_value in saved_data.items():
        target_sheet[cell_location] = cell_value

    if auto_filter:
        target_sheet = add_column_filter(target_sheet)

    if auto_column_width:
        auto_set_column_width(target_sheet)

    output_io = BytesIO()
    target_workbook.save(output_io)
    output_io.seek(0)

    smb_store_remote_file_by_obj(username, password, server_ip, server_name, share_name, remote_file_path, output_io, port)


def append_flexible_dataframe_into_excel(username, password, server_ip, server_name, share_name, remote_file_path, sheet_name, saved_data, column_name_dict,
                                         port=445, auto_filter=False, auto_column_width=False):
    """ This function is used to append data in DataFrame format to an Excel file.

    Args:

        username(str): This is the username
        password(str): This is the password
        server_ip(str): This is the ip address of the public folder, which can be same as server name
        server_name(str):This is the server name (url), e.g. szh0fs06.apac.bosch.com
        share_name(str): This is the share_name of public folder, e.g. GS_ACC_CN$
        remote_file_path(str): This is the public file path that file will be saved under share name folder
        port(int): This is the port number of the server name
        sheet_name(str): The name of the sheet.
        saved_data(pd.Dataframe): The data to be saved.
        column_name_dict(dict): The dictionary of column name and column index. e.g. {'name':'A', 'age':'C'}
        auto_filter(bool): Whether to add column filter.
        auto_column_width(bool): Whether to auto set column width.
    """
    file_obj = smb_load_file_obj(username, password, server_ip, server_name, share_name, remote_file_path, port)
    target_workbook = load_workbook(file_obj)

    target_sheet = target_workbook[sheet_name]
    max_row_number = target_sheet.max_row

    for column_name, column_index in column_name_dict.items():
        start_row_number = max_row_number + 1
        for column_value in saved_data[column_name]:
            target_sheet[f'{column_index}{start_row_number}'] = column_value
            start_row_number += 1

    if auto_filter:
        target_sheet = add_column_filter(target_sheet)

    if auto_column_width:
        auto_set_column_width(target_sheet)

    output_io = BytesIO()
    target_workbook.save(output_io)
    output_io.seek(0)

    smb_store_remote_file_by_obj(username, password, server_ip, server_name, share_name, remote_file_path, output_io, port)
