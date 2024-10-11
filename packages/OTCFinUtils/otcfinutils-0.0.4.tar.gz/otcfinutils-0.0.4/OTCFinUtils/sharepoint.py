from openpyxl.styles import PatternFill
from openpyxl import load_workbook
from OTCFinUtils.security import get_graph_token
from io import BytesIO
import pandas as pd
import requests
import base64


# TODO - review and refactor code


def load_document(dataverse_url: str, token: str, document_path: str, drive: str) -> BytesIO | None:
    site_id, drive_id = get_sp_connection_details(dataverse_url, token, drive)
    access_token = get_graph_token()

    # Now get the file from graph api
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Content-Type": "application/json",
    }
    url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives/{drive_id}"
    url += f"/root:/{document_path}?select=id,@microsoft.graph.downloadUrl"

    response = requests.get(url, headers=headers)

    if response.status_code != 200:
        print("Failed to fetch JSON data:", response.status_code)
        return

    # Get the JSON data from the response
    json_data = response.json()

    # Assuming "@microsoft.graph.downloadUrl" is the key where the URL is stored
    download_url = json_data.get("@microsoft.graph.downloadUrl")
    global _graphitemid
    _graphitemid = json_data.get("id")

    if not download_url:
        print("Download URL not found in JSON data")
        return

    # Send a GET request to the download URL
    download_response = requests.get(download_url)

    # Check if the download request was successful
    if download_response.status_code != 200:
        print("Failed to download file:", download_response.status_code)
        return

    # Get the content as a bytes object
    content = download_response.content

    return BytesIO(content)


# TODO:
# - Remove dependence on Dataverse system variable, use key vault instead
# - Use DVHandler object, instead of passing the url and token

def get_sp_connection_details(dataverse_url: str, token: str, drive: str = "account") -> tuple[str, str]:
    dataverse_url = f"{dataverse_url}/api/data/v9.2/new_systemvariableses?"
    dataverse_url += f"$select=new_value&"
    dataverse_url += f"$filter=%20new_name%20eq%20%27HTTP_SHAREPOINT_CONNECTION%27"
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}

    response = requests.get(dataverse_url, headers=headers)
    if response.status_code == 200:
        data = response.json()

        connection_details_url = ""

        if "value" in data and len(data["value"]) > 0:
            connection_details_url = data["value"][0]["new_value"]
            print("Connection Details URL:", connection_details_url)
        else:
            print("No data found in the response.")
        
        response = requests.post(connection_details_url, json={"drive": drive})

    if response.status_code != 200:
        raise RuntimeError("Failed to get sharepoint connection details")

    site_id = response.json().get("siteid")
    drive_id = response.json().get("driveid")

    return site_id, drive_id


def create_document_in_sharepoint(
    df: pd.DataFrame,
    dataverse: str,
    token: str,
    document_path: str,
    file_name: str,
    sheet: str,
):
    document_path = document_path.replace("account/", "")
    access_token = get_graph_token()
    site_id, drive_id = get_sp_connection_details(dataverse, token)

    url = f"https://graph.microsoft.com/v1.0/sites"
    url += f"/{site_id}/drives/{drive_id}/root:"
    url += f"/{document_path}/{file_name}:/content"

    headers = {
        "Authorization": f"Bearer {access_token}",
        "Content-Type": "application/octet-stream",
    }

    with BytesIO() as output:
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet, index=False)
        output.seek(0)
        print(url)
        response = requests.put(url, output, headers=headers)
        print(response.status_code)

    if response.status_code not in [201, 200]:
        print("Failed to upload file:", response.status_code)

    return response.status_code


def update_excel(excel_content: bytes, new_json_rows: list) -> str:
    df = pd.read_excel(pd.ExcelFile(BytesIO(excel_content)))

    # If DataFrame is empty, create columns using keys from the first row of new_json_rows
    if df.empty:
        df = pd.DataFrame(columns=new_json_rows[0].keys())

    # Convert each JSON row into a DataFrame and concatenate them with the existing DataFrame
    for row_data in new_json_rows:
        row_df = pd.DataFrame([row_data])
        df = pd.concat([df, row_df], ignore_index=True)

    # Convert DataFrame to Excel format and return as bytes
    output = BytesIO()
    df.to_excel(output, index=False)

    # Load the workbook
    output.seek(0)
    workbook = load_workbook(output)
    sheet = workbook.active

    if sheet is None:
        raise RuntimeError("Value of 'sheet' cannot be None.")

    # Get the range of new rows in the sheet
    start_row = len(df) - len(new_json_rows) + 2
    end_row = len(df) + 2

    # Apply yellow fill only to the range of new rows
    for row in range(start_row, end_row):
        for column in range(1, len(df.columns) + 1):
            cell = sheet.cell(row=row, column=column)
            cell.fill = PatternFill(
                start_color="FFFF00", end_color="FFFF00", fill_type="solid"
            )

    # Adjust column widths to fit the text
    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column if cell.value is not None]
        if column:
            max_length = max(len(str(cell.value)) for cell in column)
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column[0].column_letter].width = adjusted_width

    # Save the workbook to BytesIO
    output.seek(0)
    workbook.save(output)

    updated_excel_content = output.getvalue()

    # Encode Excel content to base64
    updated_excel_base64 = base64.b64encode(updated_excel_content).decode("utf-8")

    return updated_excel_base64
