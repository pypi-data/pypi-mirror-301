import numpy as np
from .data_writer import DataWriter
from openpyxl import Workbook, load_workbook


class XLSXWriter(DataWriter):

    def __init__(self, file: str = None) -> None:
        super().__init__(file)

    def write(self, data: np.ndarray, col: int = 1, file: str = None):
        if col < 1:
            raise ValueError("Column value must be greater than 0.")
        if self._file is None and file is None:
            raise ValueError("File has not been assigned or provided.")
        else:
            self._file = self._file if self._file is not None else file

        self._resource_workbook()
        sheet = self._wb.active

        for i in range(data.size):
            sheet.cell(row=i + 1, column=col, value=data[i])

        self.save()

    def save(self):
        if self._wb is None:
            ValueError("Workbook is None, nothing to write")
        self._wb.save(self._resources / self._file)

    def _resource_workbook(self):
        try:
            self._wb = load_workbook(self._resources / self._file)
        except:
            self._wb = Workbook()
