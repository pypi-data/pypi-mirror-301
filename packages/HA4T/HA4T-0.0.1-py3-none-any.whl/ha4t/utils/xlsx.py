# -*- coding: utf-8 -*-
# @时间       : 2024/8/19 10:52
# @作者       : caishilong
# @文件名      : xlsx.py
# @项目名      : pc-test
# @Software   : PyCharm
import openpyxl


class Excel:

    def __init__(self, file_path):
        """
        Excel类，用于操作Excel文件

        :param file_path:
        """
        self.file_path = file_path
        if not self.file_path.endswith('.xlsx'):
            self.file_path += '.xlsx'
        self.workbook = openpyxl.Workbook()

    def get_default_sheet(self):
        # 获取默认工作表
        return Sheet(self.workbook, self.workbook.active.title, self.file_path)

    def create_sheet(self, sheet_name):
        # 创建新的工作表
        sheet = self.workbook.create_sheet(title=sheet_name)
        return Sheet(self.workbook, sheet_name, self.file_path)

    def save(self):
        self.workbook.save(self.file_path)

    def __del__(self):
        self.save()


class Sheet:

    def __init__(self, workbook: openpyxl.Workbook, sheet_name: str, file_path: str):
        """
        Sheet类，用于操作Excel工作表
        :param workbook:
        :param sheet_name:
        :param file_path:
        """
        self.file_path = file_path
        self.workbook = workbook

        # 检查工作表是否已经存在，如果不存在则创建
        if sheet_name not in workbook.sheetnames:
            self.sheet = workbook.create_sheet(title=sheet_name)
        else:
            self.sheet = workbook[sheet_name]

    def rename_sheet(self, new_name: str):
        # 重命名工作表
        self.sheet.title = new_name

    def add_row(self, row_data: list|tuple):
        self.sheet.append(row_data)
        self.save()

    def save(self):
        # 确保保存使用的文件名是正确的
        self.workbook.save(self.file_path)


if __name__ == '__main__':
    excel = Excel('test.xlsx')
    sheet1 = excel.get_default_sheet()
    sheet1.add_row(['name', 'age', 'gender'])
    sheet1.add_row(['caishilong', 25, 'Male'])
