#!/usr/bin/env python
"""weather
"""
import pathlib
import urllib
import urllib.parse
import functools
import hashlib
import openpyxl
import requests
from .. import config


def download_district_id_data(filepath):
    """download district_id data
    """
    filepath = pathlib.Path(filepath).expanduser()
    url = config.get(
        'yycli.commands.weather.district_id_reference_download_url')
    if not url:
        url = ('https://mapopen-website-wiki.bj.bcebos.com'
               '/cityList/weather_district_id.csv')

    filepath.parent.mkdir(parents=True, exist_ok=True)

    with requests.get(url, timeout=5, stream=True) as resp:
        resp.raise_for_status()
        with open(filepath, 'wb') as fout:
            for chunk in resp.iter_content(chunk_size=8192):
                fout.write(chunk)


@functools.cache
def weather_district_data():
    """load weather_district_id.csv data
    """

    filepath = config.get('yycli.commands.weather.district_id_reference_path')
    if not filepath:
        filepath = pathlib.Path('~/.config/yycli/data/weather_district_id.csv')
    else:
        filepath = pathlib.Path(filepath)

    if not filepath.expanduser().is_file():
        download_district_id_data(filepath.as_posix())

    with open(filepath.expanduser(), 'r', encoding='utf-8') as fin:
        sheet = list(map(lambda x: x.strip().split(','), fin.readlines()))
        weather_data = sheet[1:]
        return weather_data


def download_weather_phenomenon_data(filepath):
    """download weather_phenomenon data
    """
    filepath = pathlib.Path(filepath).expanduser()
    url = config.get(
        'yycli.commands.weather.phenomenon_reference_download_url')
    if not url:
        url = ('https://mapopen-website-wiki.cdn.bcebos.com/cityList/'
               '%E7%99%BE%E5%BA%A6%E5%9C%B0%E5%9B%BE%E5%A4%A9%E6'
               '%B0%94%E5%8F%96%E5%80%BC%E5%AF%B9%E7%85%A7%E8%A1'
               '%A8(0410).xlsx')

    filepath.parent.mkdir(parents=True, exist_ok=True)

    with requests.get(url, timeout=5, stream=True) as resp:
        resp.raise_for_status()
        with open(filepath, 'wb') as fout:
            for chunk in resp.iter_content(chunk_size=8192):
                fout.write(chunk)


@functools.cache
def weather_phenomenon_data():
    """load weather_phenomenon.txt data
    """
    filepath = config.get('yycli.commands.weather.phenomenon_reference_path')
    if not filepath:
        filepath = pathlib.Path(
            '~/.config/yycli/data/weather_phenomenon_reference.xlsx')
    else:
        filepath = pathlib.Path(filepath)

    if not filepath.expanduser().is_file():
        download_weather_phenomenon_data(filepath.as_posix())

    workbook = openpyxl.load_workbook(filepath.expanduser())
    worksheet = workbook.get_sheet_by_name('天气现象')
    rows = worksheet.iter_rows(2, worksheet.max_row, 1, worksheet.max_column)
    sheet = list(map(lambda x: [cell.value for cell in x], rows))
    return sheet


def querystring(params):
    """get query string
    """
    return '&'.join(map(lambda x: f'{x[0]}={x[1]}', params.items()))


def sign(access_key, secure_key, apipath: str, params=None):
    """get signature
    """
    path_query_str = apipath
    if not params:
        query = {'ak': access_key}
    else:
        query = {**params, 'ak': access_key}

    path_query_str = f'{apipath}?{querystring(query)}'

    encode_str = urllib.parse.quote(path_query_str, safe='/:=&?#+!$,;\'@()*[]')
    signature = hashlib.md5(
        urllib.parse.quote_plus(
            (encode_str + secure_key)).encode()).hexdigest()
    return signature


@functools.cache
def lookup_weather_district_by_text(district):
    """search district in weather_district_data by district text
    """
    for row in weather_district_data():
        if row[5].startswith(district):
            return row
    return None


@functools.cache
def lookup_weather_district_by_id(district_id):
    """search district in weather_district_data by district id
    """
    for row in weather_district_data():
        if row[4] == district_id:
            return row
    return None


def args_parser(parser):
    """args parser
    """
    parser.add_argument('-f',
                        '--from-file',
                        type=str,
                        dest='from_file',
                        help='read district id list from file',
                        default=None)
    parser.add_argument('district_id', nargs='*', help='district id')


def query_weather_by_district_id(district_id):
    """query weather by district id
    """
    access_key = config.get('yycli.commands.weather.access_key')
    secure_key = config.get('yycli.commands.weather.secure_key')

    req_params = {
        'district_id': district_id,
        'data_type': 'all',
    }
    entrypoint = 'https://api.map.baidu.com'
    apipath = '/weather/v1/'
    signature = sign(access_key, secure_key, apipath, req_params)
    params = {**req_params, 'ak': access_key, 'sn': signature}
    req = requests.get(f'{entrypoint}{apipath}'
                       f'?{querystring(params)}',
                       timeout=5)
    res = req.json()
    ret = {
        'area': res['result']['location']['name'],
        'phenomenon': res['result']['now']['text'],
        'temperature': res['result']['now']['temp'],
        'wind_directory': res['result']['now']['wind_dir'],
        'wind_class': res['result']['now']['wind_class'],
    }
    return ret


def resolve_district_id(district_id):
    """resolve district id
    """
    if district_id.isnumeric():
        # check if district_id valid
        if not lookup_weather_district_by_id(district_id):
            # invalid district id, ignore
            return None
        return int(district_id)
    row = lookup_weather_district_by_text(district_id)
    if row:
        return row[4]
    return None


def weather(args):
    """weather
    """
    # use Beijing 110100 as default
    default_district_id = 110100
    district_id_list = []
    if args.from_file and pathlib.Path(args.from_file).expanduser().exists():
        with open(pathlib.Path(args.from_file).expanduser(),
                  'r',
                  encoding='utf-8') as fin:
            for line in fin.readlines():
                district_id = resolve_district_id(line.strip())
                district_id_list.append(district_id)
    else:
        for district_id in args.district_id:
            district_id = resolve_district_id(district_id)
            if district_id:
                district_id_list.append(district_id)

    if not district_id_list:
        district_id_list.append(default_district_id)

    for district_id in district_id_list:
        ret = query_weather_by_district_id(district_id)

        format_string = config.get('yycli.commands.weather.format_string')
        if not format_string:
            format_string = ('%(area)s %(phenomenon)s %(temperature)sºC'
                             ' %(wind_directory)s/%(wind_class)s')

        print(format_string % ret)
